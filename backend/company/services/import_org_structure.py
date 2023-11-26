import openpyxl
from typing import Dict, List, Any, BinaryIO

from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.dialects.postgresql import insert as pg_insert

from company.models.db import Employee, Assignment, StaffUnit, Division, Acting
from database import async_session_maker


def parse_xlsx(file_path: BinaryIO) -> Dict[str, Worksheet]:
    """
    Функция для подготовки к парсингу файла XLSX.
    Здесь указываются листы, которые необходимо обработать в этом файле.

    :param file_path: Передаваемый файл XLSX.
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet_division = workbook["Подразделения"]
    sheet_employees = workbook["Сотрудники"]
    sheet_staff_units = workbook["Штатные единицы"]
    sheet_assignment = workbook["Назначения"]
    sheet_acting = workbook["ВрИО"]
    workbook.close()
    return {
        'sheet_division': sheet_division,
        'sheet_employees': sheet_employees,
        'sheet_staff_units': sheet_staff_units,
        'sheet_assignment': sheet_assignment,
        'sheet_acting': sheet_acting,
    }


def parse_sheet(sheet: Worksheet, max_columns: int) -> List[Dict]:
    """
    Функция для парсинга листа таблицы.
    Преобразует данные в словари с ключами, соответствующими заголовкам в листе.

    :param sheet: Лист таблицы XLSX.
    :param max_columns: Максимальное количество колонок, которые необходимо учесть при чтении листа.
    """
    rows = []
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1, max_col=max_columns))]

    if len(headers) != max_columns:
        raise Exception

    for row in sheet.iter_rows(min_row=2, max_col=max_columns):
        if row[0].value is None:
            return rows

        rows.append(dict(zip(headers, [cell.value for cell in row])))
    return rows


async def insert_or_update_db(data: List[Dict[str, Any]], model) -> None:
    """
    Функция добавления новых и обновления уже существующих строк в БД.

    :param data: Список словарей, соответствующий строкам таблицы структуры организации.
    :param model: Модель базы данных, в которую будет доавлена строка.
    """
    async with async_session_maker() as session:
        insert_stmt = pg_insert(model).values(data)
        update_columns = {col.name: col for col in insert_stmt.excluded if col.name not in ('id')}
        update_statement = insert_stmt.on_conflict_do_update(index_elements=['id'], set_=update_columns)

        await session.execute(update_statement)
        await session.commit()


async def import_org_structure(filename: BinaryIO) -> dict[str, list[dict]]:
    """
    Функция для вызова в API для импорта структуры организации.

    В функции указываются, какие листы XLSX будут обрабатываться,
    а к каждому листу указывается ограничение количества колонок.

    :param filename: Передаваемый файл формата XLSX, содержащий структуру организации
    """
    parsed_table = parse_xlsx(file_path=filename)
    employees = parse_sheet(sheet=parsed_table['sheet_employees'], max_columns=7)
    divisions = parse_sheet(sheet=parsed_table['sheet_division'], max_columns=5)
    staff_units = parse_sheet(sheet=parsed_table['sheet_staff_units'], max_columns=3)
    assignments = parse_sheet(sheet=parsed_table['sheet_assignment'], max_columns=6)
    acting = parse_sheet(sheet=parsed_table['sheet_acting'], max_columns=6)

    await insert_or_update_db(data=employees, model=Employee)
    await insert_or_update_db(data=divisions, model=Division)
    await insert_or_update_db(data=staff_units, model=StaffUnit)
    await insert_or_update_db(data=assignments, model=Assignment)
    await insert_or_update_db(data=acting, model=Acting)

    return {
        'employees': employees,
        'divisions': divisions,
        'staff_units': staff_units,
        'assignments': assignments,
        'acting': acting,
    }

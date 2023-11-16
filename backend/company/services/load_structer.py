import openpyxl

# Укажите путь к вашему файлу XLSX
file_path = ""

# Открываем файл
workbook = openpyxl.load_workbook(file_path)

# Парсинг страницы "Подразделение"
sheet_department = workbook["Подразделение"]

# Парсинг страницы "Сотрудники"
sheet_employees = workbook["Сотрудники"]

# Парсинг страницы "Назначения"
sheet_appointments = workbook["Назначения"]

# Парсинг страницы "Штатные единицы"
sheet_positions = workbook["Штатные единицы"]

# Парсинг страницы "ВрИО"
sheet_temporary_heads = workbook["ВрИО"]

# Пример чтения данных с каждой страницы
# На странице "Подразделение"
for row in sheet_department.iter_rows():
    for cell in row:
        print(cell.value, end="\t")
    print()

# На странице "Сотрудники"
for row in sheet_employees.iter_rows():
    for cell in row:
        print(cell.value, end="\t")
    print()

# На странице "Назначения"
for row in sheet_appointments.iter_rows():
    for cell in row:
        print(cell.value, end="\t")
    print()

# На странице "Штатные единицы"
for row in sheet_positions.iter_rows():
    for cell in row:
        print(cell.value, end="\t")
    print()

# На странице "ВрИО"
for row in sheet_temporary_heads.iter_rows():
    for cell in row:
        print(cell.value, end="\t")
    print()

# Закрываем файл после использования
workbook.close()


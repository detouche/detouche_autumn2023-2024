import io

import openpyxl
from fastapi import APIRouter, Response, HTTPException
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill

from auth.services.users_api import get_employee_info
from company.repository.company import DivisionRepository, EmployeeRepository
from docs.repository.docs import DocumentRepository, CourseRepository

document_report_router = APIRouter(prefix='/document-report', tags=['document-reports'])

document_repository = DocumentRepository()
course_repository = CourseRepository()
division_repository = DivisionRepository()
employee_repository = EmployeeRepository()


@document_report_router.post('/courses')
async def export_courses_report():
    documents = await document_repository.find_all()

    if not documents:
        raise HTTPException(status_code=404, detail={
            'code': 'DOCUMENTS_NOT_FOUND',
            'reason': 'There is no documents'
        })

    data = []
    for document in documents:
        course = await course_repository.find_one(document.course_id)
        member = await get_employee_info(await employee_repository.find_one(document.members_id[0]))

        formatted_members = []
        for member_id in document.members_id:
            employee = await employee_repository.find_one(member_id)
            formatted_members.append(f'{employee.surname} {employee.name} {employee.patronymic}')

        data.append({
            'education_center': course.education_center,
            'title': course.title,
            'type': course.type,
            'category': course.category,
            'division': member['division']['name'],
            'employees_name': ';\n'.join(formatted_members),
            'cost_all': len(document.members_id) * course.cost,
            'cost_one': course.cost,
            'date': f'{course.start_date} - {course.end_date}',
        })

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append([
        "Учебный центр",
        "Название курса",
        "Тип",
        "Направление обучения",
        "Подразделение",
        "ФИО сотрудников",
        "Сумма затрат",
        "Стоимость на 1 человека",
        "Даты"
    ])

    for row in data:
        ws.append([row["education_center"], row["title"], row["type"], row["category"], row["division"],
                   row["employees_name"], row["cost_all"], row["cost_one"], row["date"]])

    style = NamedStyle(name='Main')
    style.alignment.wrap_text = True
    style.alignment.shrinkToFit = True
    wb.add_named_style(style)

    headings_style = NamedStyle(name='Headings')
    headings_style.alignment.horizontal = 'center'
    headings_style.alignment.vertical = 'center'
    headings_style.font = Font(bold=True)
    headings_style.border = Border(left=Side(style='thin'),
                                   right=Side(style='thin'),
                                   top=Side(style='thin'),
                                   bottom=Side(style='thin'))
    headings_style.alignment.wrap_text = True
    headings_style.alignment.shrinkToFit = True
    headings_style.fill = PatternFill(start_color='D9D9D9',
                                      end_color='D9D9D9',
                                      fill_type='solid')
    wb.add_named_style(headings_style)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 13
    ws.column_dimensions['I'].width = 21

    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        for cell in col:
            cell.style = 'Main'

    for cell in ws['1:1']:
        cell.style = 'Headings'

    buffer = io.BytesIO()
    wb.save(buffer)

    buffer.seek(0)

    return Response(buffer.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={'Content-Disposition': 'attachment; filename="Documents Report.xlsx"'})
